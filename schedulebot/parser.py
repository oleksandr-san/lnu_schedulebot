import json
import os
import re
from datetime import date, datetime, time
from typing import Any, Iterator, List, Optional

import openpyxl

from schedulebot.schedule import (
    ScheduleTemplate,
    ScheduleTemplateItem,
    WeekParity,
    unstructure_template,
)

DATE_PATTERN = re.compile(r'\b\d{2}\.\d{2}\b')


def parse_weekday(value: Any) -> int:
    if isinstance(value, str):
        value = value.lower().split(' ')[0].replace("'", '').replace('"', '')
        weekday = {
            'понеділок': 0,
            'вівторок': 1,
            'середа': 2,
            'четвер': 3,
            'пятниця': 4,
            'субота': 5,
            'неділя': 6,
        }.get(value, 0)
        return weekday
    elif isinstance(value, int):
        return value % 7
    else:
        raise TypeError('unsupported weekday')


def parse_time(time_str: str) -> time:
    hour, minute = map(int, time_str.split(':'))
    return time(hour, minute)


def parse_dates_from_text(text: str) -> List[date]:
    found_dates = DATE_PATTERN.findall(text)
    current_year = datetime.now().year
    return [
        date(current_year, int(d.split('.')[1]), int(d.split('.')[0]))
        for d in found_dates
    ]


def parse_schedule_template(workbook: openpyxl.Workbook) -> Iterator[ScheduleTemplate]:  # noqa: C901, PLR0912, PLR0914, PLR0915
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        items = []

        merged_ranges_by_start = {}
        for merged_range in sheet.merged_cells.ranges:
            start_cell = getattr(merged_range, 'start_cell', None)
            if start_cell:
                merged_ranges_by_start[start_cell.coordinate] = merged_range

        title_row = next(sheet.iter_rows(min_row=2, max_row=3), None)
        if not title_row or not title_row[0].value:
            continue
        title = str(title_row[0].value).strip()

        group_by_column = {}
        cur_weekday: Optional[int] = None
        cur_slot: Optional[int] = None
        cur_time: Optional[time] = None
        cur_time_cnt = 0

        for i, row in enumerate(sheet.iter_rows(min_row=3)):
            header_start = ['День', 'Пара', 'Час']
            if len(row) < len(header_start):
                continue

            first3 = list(
                filter(
                    None,
                    [cell.value and str(cell.value).strip() for cell in row[:3]],
                )
            )
            if first3 == header_start:
                for cell in row[3:]:
                    if not cell.value:
                        continue
                    group = str(cell.value).strip()
                    group_by_column[cell.col_idx] = group
                continue

            if raw_weekday := row[0].value:
                cur_weekday = parse_weekday(raw_weekday)
            if raw_slot := row[1].value:
                cur_slot = int(str(raw_slot))
            if raw_time := row[2].value:
                cur_time = time.fromisoformat(str(raw_time))
                cur_time_cnt = 0
            elif cur_time_cnt < 1:
                cur_time_cnt += 1
            else:
                cur_time = None
                cur_time_cnt = 0

            if cur_weekday is None or cur_slot is None or cur_time is None:
                continue

            print(f'Row {i}:', cur_weekday, cur_slot, cur_time)
            for col in row[3:]:
                if col.value is None:
                    continue

                try:
                    group = group_by_column[col.col_idx]
                except (AttributeError, KeyError):
                    continue

                mr = merged_ranges_by_start.get(col.coordinate)

                if mr is not None and mr.min_row != mr.max_row:
                    week_parity = None
                elif col.row % 2 == 0:
                    week_parity = WeekParity.ODD
                else:
                    week_parity = WeekParity.EVEN

                if mr is not None:
                    groups = [
                        group_by_column[col_idx]
                        for col_idx in range(mr.min_col, mr.max_col + 1)
                    ]
                else:
                    groups = [group]

                print(col.coordinate, mr, col.value, week_parity, groups)
                name = str(col.value).strip()
                item = ScheduleTemplateItem(
                    slot=cur_slot,
                    name=name,
                    weekday=cur_weekday,
                    start=cur_time,
                    groups=groups,
                    week_parity=week_parity,
                    dates=parse_dates_from_text(name) or None,
                )
                items.append(item)

        if items:
            template = ScheduleTemplate(name=sheet_name, title=title, items=items)
            yield template


if __name__ == '__main__':
    file_path = './data/ROZKLAD_MV_1_SEMESTR_2024_30_09.xlsx'
    output_path = './schedulebot/data/ROZKLAD_MV_1_SEMESTR_2024_30_09'

    os.makedirs(output_path, exist_ok=True)

    workbook = openpyxl.load_workbook(file_path)
    schedule_templates = parse_schedule_template(workbook)

    for template in schedule_templates:
        template_filename = template.name.lower().replace(' ', '_') + '.json'
        template_path = os.path.join(output_path, template_filename)

        with open(template_path, 'w', encoding='utf-8') as fp:
            json.dump(
                unstructure_template(template),
                fp,
                indent=4,
                default=str,
                ensure_ascii=False,
            )
